from dateutil.relativedelta import relativedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from datetime import date, timedelta

from django.core.exceptions import ValidationError
from .forms import RefillForm, UploadExcelForm
from .models import Refill, Facility
import pandas as pd
import openpyxl
from django.http import HttpResponse
from django.conf import settings
from django.db.models import F, Q
from django.core.paginator import Paginator
from openpyxl import Workbook
from datetime import datetime

from django.contrib import messages
from openpyxl.styles import Font
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from .models import FacilityUser
from django.contrib.auth.decorators import user_passes_test

from django.core.exceptions import ObjectDoesNotExist

# views.py






def signin_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('pw')

        user = authenticate(username=username, password=password)

        if user is not None:

            # ✅ SUPERUSER BYPASS
            if not user.is_superuser:
                if not hasattr(user, "facility_membership"):
                    messages.error(request, "User not assigned to any facility.")
                    return redirect('login')

            login(request, user)
            messages.success(request, 'Login successful!')
            return redirect('dashboard')

        messages.error(request, "Invalid credentials")

    return render(request, 'signin.html')





def logout_view(request):
    logout(request)
    messages.success(request, ' Logout successfully')
    return redirect('login')




# ================= USER / FACILITY HELPERS =================

def get_user_facility(user):
    if user.is_superuser:
        return None

    if hasattr(user, "facility_membership"):
        return user.facility_membership.facility

    return None




def is_admin(user):
    if hasattr(user, "facility_membership"):
        return user.facility_membership.role == "admin"
    return False





def attach_vl_status(refill):
    if not refill:
        return None, None

    eligible = refill.is_vl_eligible_program
    status = "Eligible" if eligible else "Not Eligible"

    return eligible, status




def user_refill_queryset(user):
    qs = Refill.objects.select_related("facility")

    if not user.is_superuser:
        qs = qs.filter(facility__memberships__user=user)

    return qs






def admin_required(view_func):
    def check(user):
        return (
            user.is_authenticated and
            hasattr(user, "facility_membership") and
            user.facility_membership.role == "admin"
        )

    return user_passes_test(check)(view_func)






def get_quarter(today):
    if today.month in [1, 2, 3]:
        return date(today.year, 1, 1), date(today.year, 3, 31)

    elif today.month in [4, 5, 6]:
        return date(today.year, 4, 1), date(today.year, 6, 30)

    elif today.month in [7, 8, 9]:
        return date(today.year, 7, 1), date(today.year, 9, 30)

    return date(today.year, 10, 1), date(today.year, 12, 31)


# -----------------------
# Excel import function
# -----------------------


VALID_REFILL_MONTHS = [0.5, 1, 2, 2.8, 3, 4, 5, 6]


def clean_int(value):
    if pd.isnull(value):
        return None

    value = str(value).strip().replace(",", "")

    if value.lower() in ["failed", "n/a", "na", "not done", "--", ""]:
        return None

    try:
        return int(float(value))
    except Exception:
        return None


def import_refills_from_excel(file):

    MAX_FILE_SIZE = 5 * 1024 * 1024 * 1024  # 5 GB

    if file.size > MAX_FILE_SIZE:
        raise ValidationError("File size exceeds 50 MB limit")

    file.seek(0)

    df = pd.read_excel(file)

    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '').str.lower()

    required_columns_map = {
        'unique id': 'unique id',
        'last pickup date (yyyy-mm-dd)': 'last pickup date (yyyy-mm-dd)',
        'months of arv refill': 'months of arv refill',
        'current art regimen': 'current art regimen',
        'case manager': 'case manager',
        'sex': 'sex',
        'current art status': 'current art status',
        'facility name': 'facility name',

        'art start date (yyyy-mm-dd)': 'art start date (yyyy-mm-dd)',
        'date of viral load sample collection (yyyy-mm-dd)': 'date of viral load sample collection (yyyy-mm-dd)',
        'current viral load (c/ml)': 'current viral load (c/ml)',

        'date of tpt start (yyyy-mm-dd)': 'date of tpt start (yyyy-mm-dd)',
        'tpt completion date (yyyy-mm-dd)': 'tpt completion date (yyyy-mm-dd)',

        'date of commencement of eac (yyyy-mm-dd)': 'date of commencement of eac (yyyy-mm-dd)',
        'number of eac sessions completed': 'number of eac sessions completed',

        'age': 'age',
        'date of tb screening (yyyy-mm-dd)': 'date of tb screening (yyyy-mm-dd)',
        'tb screening type': 'tb screening type',
        'tb status': 'tb status',
        'date of tb sample collection (yyyy-mm-dd)': 'date of tb sample collection (yyyy-mm-dd)',
        'date of tb diagnostic result received (yyyy-mm-dd)': 'date of tb diagnostic result received (yyyy-mm-dd)',
        'tb diagnostic result': 'tb diagnostic result',
    }

    missing_columns = [col for col in required_columns_map.values() if col not in df.columns]

    if missing_columns:
        raise ValidationError(f"Missing column(s): {', '.join(missing_columns)}")

    df = df[df['current art status'].isin(['Active', 'Active Restart'])]

    if df.empty:
        raise ValidationError("No Active or Active Restart patients found.")

    df['facility name'] = df['facility name'].astype(str).str.strip()

    facilities = {
        f.name.strip(): f for f in Facility.objects.filter(
            name__in=df['facility name'].unique()
        )
    }

    missing_facilities = set(df['facility name'].unique()) - set(facilities.keys())

    if missing_facilities:
        raise ValidationError(
            f"Facilities not in system: {', '.join(missing_facilities)}"
        )

    validated_rows = []

    for _, row in df.iterrows():

        unique_id = row['unique id']

        try:
            last_pickup_date = pd.to_datetime(
                row['last pickup date (yyyy-mm-dd)']
            ).date()
        except Exception:
            raise ValidationError(f"Invalid Last Pickup Date for {unique_id}")

        try:
            months = float(row['months of arv refill'])
        except Exception:
            raise ValidationError(
                f"Invalid Months of ARV Refill for {unique_id}"
            )

        if months not in VALID_REFILL_MONTHS:
            raise ValidationError(
                f"Invalid refill months {months} for {unique_id}"
            )

        facility_obj = facilities[row['facility name']]

        next_appointment = last_pickup_date + timedelta(days=months * 30)

        # ================= OPTIONAL FIELDS =================

        art_start_date = pd.to_datetime(
            row['art start date (yyyy-mm-dd)']
        ).date() if pd.notnull(row['art start date (yyyy-mm-dd)']) else None

        vl_sample_collection_date = pd.to_datetime(
            row['date of viral load sample collection (yyyy-mm-dd)']
        ).date() if pd.notnull(row['date of viral load sample collection (yyyy-mm-dd)']) else None

        vl_result = clean_int(
            row['current viral load (c/ml)']
        ) if pd.notnull(row['current viral load (c/ml)']) else None

        tpt_start_date = pd.to_datetime(
            row['date of tpt start (yyyy-mm-dd)']
        ).date() if pd.notnull(row['date of tpt start (yyyy-mm-dd)']) else None

        tpt_completion_date = pd.to_datetime(
            row['tpt completion date (yyyy-mm-dd)']
        ).date() if pd.notnull(row['tpt completion date (yyyy-mm-dd)']) else None

        tpt_expected_completion = (
            tpt_start_date + timedelta(days=180)
            if tpt_start_date else None
        )

        eac_start_date = pd.to_datetime(
            row['date of commencement of eac (yyyy-mm-dd)']
        ).date() if pd.notnull(row['date of commencement of eac (yyyy-mm-dd)']) else None

        eac_sessions_completed = clean_int(
            row['number of eac sessions completed']
        ) if pd.notnull(row['number of eac sessions completed']) else 0

        # ================= NEW TB FIELDS =================

        age = clean_int(row['age']) if pd.notnull(row['age']) else None

        tb_screening_date = pd.to_datetime(
            row['date of tb screening (yyyy-mm-dd)']
        ).date() if pd.notnull(row['date of tb screening (yyyy-mm-dd)']) else None

        tb_screening_type = str(row['tb screening type']).strip() if pd.notnull(
            row['tb screening type']) else None

        tb_status = str(row['tb status']).strip() if pd.notnull(
            row['tb status']) else None

        tb_sample_collection_date = pd.to_datetime(
            row['date of tb sample collection (yyyy-mm-dd)']
        ).date() if pd.notnull(
            row['date of tb sample collection (yyyy-mm-dd)']
        ) else None

        tb_result_received_date = pd.to_datetime(
            row['date of tb diagnostic result received (yyyy-mm-dd)']
        ).date() if pd.notnull(
            row['date of tb diagnostic result received (yyyy-mm-dd)']
        ) else None

        tb_diagnostic_result = str(row['tb diagnostic result']).strip() if pd.notnull(
            row['tb diagnostic result']) else None

        validated_rows.append(

            Refill(

                facility=facility_obj,
                unique_id=unique_id,   # ✅ duplicates now allowed

                last_pickup_date=last_pickup_date,
                months_of_refill_days=months,
                next_appointment=next_appointment,

                current_regimen=str(row['current art regimen']).strip(),
                case_manager=str(row['case manager']).strip(),
                sex=str(row['sex']).strip(),

                current_art_status=row['current art status'].strip(),

                art_start_date=art_start_date,
                vl_sample_collection_date=vl_sample_collection_date,
                vl_result=vl_result,

                tpt_start_date=tpt_start_date,
                tpt_completion_date=tpt_completion_date,
                tpt_expected_completion=tpt_expected_completion,

                eac_start_date=eac_start_date,
                eac_sessions_completed=eac_sessions_completed,

                age=age,
                tb_screening_date=tb_screening_date,
                tb_screening_type=tb_screening_type,
                tb_status=tb_status,
                tb_sample_collection_date=tb_sample_collection_date,
                tb_result_received_date=tb_result_received_date,
                tb_diagnostic_result=tb_diagnostic_result,
            )
        )

    facility_ids = {obj.facility.id for obj in validated_rows}

    with transaction.atomic():

        for facility_id in facility_ids:
            Refill.objects.filter(facility_id=facility_id).delete()

        Refill.objects.bulk_create(validated_rows, batch_size=1000)

    return len(validated_rows)




def upload_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if not request.FILES:
            messages.error(request, "No file was uploaded.")
            return redirect('upload_excel')
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            if excel_file.size >  50 * 1024 * 1024 :
                messages.error(request, "File size exceeds 50 GB limit.")
                return redirect('upload_excel')
            try:
                count = import_refills_from_excel(excel_file)
                messages.success(request, f"Excel uploaded! {count} records imported.")
                return redirect('upload_excel')
            except ValidationError as e:
                messages.error(request, str(e))
                return redirect('upload_excel')
            except Exception as e:
                messages.error(request, f"Upload failed: {str(e)}")
                return redirect('upload_excel')
        else:
            messages.error(request, "Form validation failed.")
            return redirect('upload_excel')
    else:
        form = UploadExcelForm()
    return render(request, 'upload.html', {'form': form})
# ----------------------------
# EXCEL UPLOAD VIEW
# ----------------------------


def upload_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if not request.FILES:
            messages.error(request, "No file was uploaded.")
            return redirect('upload_excel')

        if form.is_valid():
            excel_file = form.cleaned_data['file']
            if excel_file.size > 1073741824:
                messages.error(request, "File size exceeds the 1GB limit.")
                return redirect('upload_excel')
            try:
                count = import_refills_from_excel(excel_file)
                messages.success(request, f"Excel uploaded successfully! {count} records imported.")
                return redirect('upload_excel')
            except ValidationError as e:
                messages.error(request, str(e))
                return redirect('upload_excel')
            except Exception as e:
                messages.error(request, f"Upload failed: {str(e)}")
                return redirect('upload_excel')
        else:
            messages.error(request, "Form validation failed.")
            return redirect('upload_excel')
    else:
        form = UploadExcelForm()

    return render(request, 'upload.html', {'form': form})

@login_required
def dashboard(request):

    today = timezone.now().date()

    # ================= FACILITY ACCESS FIX =================
    facility_id = request.GET.get("facility")
    user = request.user

    if user.is_superuser:
        facilities = Facility.objects.all().order_by("name")

        # superuser sees all OR filters by one facility
        if facility_id:
            refills = Refill.objects.filter(
                facility_id=facility_id,
                current_art_status__in=["Active", "Active Restart", "Restart"]
            )
        else:
            refills = Refill.objects.filter(
                current_art_status__in=["Active", "Active Restart", "Restart"]
            )

        selected_facility = facility_id

    else:
        facility = get_user_facility(user)
        if not facility:
            return redirect("login")

        facilities = [facility]

        refills = Refill.objects.filter(
            facility=facility,
            current_art_status__in=["Active", "Active Restart", "Restart"]
        )

        selected_facility = facility.id

    # ================= TIME WINDOWS =================
    week_end = today + timedelta(days=7)

    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    tb_start = date(2025, 4, 1)
    tb_end = date(2026, 9, 30)

    def in_tb_period(d):
        return d and tb_start <= d <= tb_end

    # ================= COUNTERS =================
    daily_expected = 0
    daily_refills = 0
    weekly_expected = 0
    monthly_expected = 0
    monthly_missed_total = 0
    iit_total = 0

    eac_count = 0
    post_eac_vl_count = 0
    ahd_count = 0

    vl_samples = 0
    vl_results = 0
    suppressed = 0

    tb_screened = 0
    tb_presumptive = 0
    tb_sample_collected = 0
    tb_result_received = 0
    tb_positive = 0
    tb_negative = 0

    # ================= LOOP =================
    for r in refills:

        next_appt = r.next_appointment
        last_pickup = r.last_pickup_date
        days_missed = r.days_missed

        if r.vl_sample_collection_date:
            vl_samples += 1

        if r.vl_result is not None:
            vl_results += 1
            if r.vl_result < 1000:
                suppressed += 1

        if r.ahd:
            ahd_count += 1

        if in_tb_period(r.eac_start_date):
            eac_count += 1

        if r.eac_status == "Post-EAC VL Due":
            post_eac_vl_count += 1

        if next_appt:

            if month_start <= next_appt <= month_end:
                monthly_expected += 1

            if today <= next_appt <= week_end:
                weekly_expected += 1

            if next_appt == today:
                daily_expected += 1

            if last_pickup == today:
                daily_refills += 1

            if days_missed > 0:
                monthly_missed_total += 1

            if r.iit_status == "IIT":
                iit_total += 1

        if in_tb_period(r.tb_screening_date):
            tb_screened += 1

        if r.tb_status == "Presumptive TB":
            tb_presumptive += 1

        if in_tb_period(r.tb_sample_collection_date):
            tb_sample_collected += 1

        if in_tb_period(r.tb_result_received_date):
            tb_result_received += 1

            if r.tb_diagnostic_result == "Positive":
                tb_positive += 1
            elif r.tb_diagnostic_result == "Negative":
                tb_negative += 1

    vl_suppression_rate = round((suppressed / vl_results) * 100, 1) if vl_results else 0

    return render(request, "dashboard.html", {

        "facilities": facilities,
        "selected_facility": selected_facility,

        # APPOINTMENTS
        "daily_expected": daily_expected,
        "daily_refills": daily_refills,
        "weekly_expected": weekly_expected,
        "monthly_expected": monthly_expected,
        "monthly_missed_total": monthly_missed_total,
        "iit_total": iit_total,

        # PROGRAM
        "eac_count": eac_count,
        "post_eac_vl_count": post_eac_vl_count,
        "ahd_count": ahd_count,

        # VL
        "vl_samples": vl_samples,
        "vl_results": vl_results,
        "vl_suppression_rate": vl_suppression_rate,
        "suppressed": suppressed,

        # TB
        "tb_start": tb_start,
        "tb_end": tb_end,
        "tb_screened": tb_screened,
        "tb_presumptive": tb_presumptive,
        "tb_sample_collected": tb_sample_collected,
        "tb_result_received": tb_result_received,
        "tb_positive": tb_positive,
        "tb_negative": tb_negative,
    })
    
    
    
    
@login_required
def refill_list(request):

    today = timezone.now().date()
    week_end = today + timedelta(days=7)

    quarter_start, quarter_end = get_quarter(today)

    selected_case_manager = request.GET.get("case_manager")
    search_unique_id = request.GET.get("search_unique_id")
    facility_id = request.GET.get("facility")

    # ================= FACILITY HANDLING =================
    if request.user.is_superuser:

        # superuser sees all facilities
        facilities = Facility.objects.all().order_by("name")

        # if facility selected → filter, else show ALL
        if facility_id:
            refills = Refill.objects.filter(facility_id=facility_id)
        else:
            refills = Refill.objects.all()

    else:
        facility = get_user_facility(request.user)
        if not facility:
            return redirect("login")

        facilities = [facility]
        refills = Refill.objects.filter(facility=facility)

    # ================= FILTERS =================
    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    if search_unique_id:
        refills = refills.filter(unique_id__icontains=search_unique_id.strip())

    # ================= CASE MANAGERS =================
    case_managers_qs = refills.exclude(
        case_manager__isnull=True
    ).exclude(
        case_manager__exact=""
    ).values_list("case_manager", flat=True).distinct()

    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm})

    # ================= ENRICH =================
    enriched = []

    for r in refills:

        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status

        r.is_overdue = r.vl_due_date and r.vl_due_date < today
        r.missed_appointment = (r.days_missed or 0) > 0

        enriched.append(r)

    # ================= SORT =================
    enriched.sort(key=lambda x: (
        not x.vl_eligible,
        x.vl_due_date or today,
        x.next_appointment or today
    ))

    # ================= PERIODS =================
    daily_qs = [r for r in enriched if r.next_appointment == today]

    weekly_qs = [
        r for r in enriched
        if r.next_appointment and today <= r.next_appointment <= week_end
    ]

    monthly_qs = [
        r for r in enriched
        if r.next_appointment and r.next_appointment.month == today.month
    ]

    periods = [
        {
            "name": "Daily",
            "page_obj": Paginator(daily_qs, 10).get_page(request.GET.get("daily_page"))
        },
        {
            "name": "Weekly",
            "page_obj": Paginator(weekly_qs, 10).get_page(request.GET.get("weekly_page"))
        },
        {
            "name": "Monthly",
            "page_obj": Paginator(monthly_qs, 10).get_page(request.GET.get("monthly_page"))
        },
    ]

    return render(request, "refill_list.html", {
        "facilities": facilities,
        "case_managers": case_managers,
        "selected_case_manager": selected_case_manager,
        "search_unique_id": search_unique_id,
        "selected_facility": facility_id,
        "periods": periods,
        "today": today,
        "quarter_start": quarter_start,
        "quarter_end": quarter_end,
        "is_admin": is_admin(request.user),
    })
# ================= EXPORT =================

@login_required
def export_refills_view(request):

    today = timezone.now().date()

    selected_case_manager = request.GET.get("case_manager")
    search_unique_id = request.GET.get("search_unique_id")

    # ================= FACILITY HANDLING =================
    if request.user.is_superuser:
        refills = Refill.objects.all()
    else:
        facility = get_user_facility(request.user)
        if not facility:
            return redirect("login")
        refills = Refill.objects.filter(facility=facility)

    # ================= FILTERS =================
    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    if search_unique_id:
        refills = refills.filter(unique_id__icontains=search_unique_id.strip())

    # ================= ENRICH =================
    enriched = []

    for r in refills:
        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status
        enriched.append(r)

    # ================= SORT =================
    enriched.sort(key=lambda x: (
        not x.vl_eligible,
        x.vl_due_date or today,
        x.next_appointment or today
    ))

    # ================= EXPORT =================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expected Refills"

    headers = [
        "Unique ID", "Age", "Sex", "Facility", "Case Manager",
        "Next Appointment", "VL Due Date", "VL Status",
        "VL Eligible", "VL Suppression", "EAC", "AHD",
        "TPT", "TB", "Tracking Date 1", "Tracking Date 2",
        "Tracking Date 3", "Tracked By", "Missed Reason",
        "Patient Discontinued", "Discontinued Reason",
        "Discontinued Date", "Returned Date", "Remark",
    ]

    ws.append(headers)

    for r in enriched:

        suppression = (
            "Suppressed (<1000)" if r.is_suppressed else
            "Not Suppressed (≥1000)" if r.vl_result is not None else
            "No Result"
        )

        discontinued = dict(
            r._meta.get_field("patient_discontinued").choices
        ).get(r.patient_discontinued, "") if r.patient_discontinued else ""

        ws.append([
            r.unique_id,
            r.age or "",
            r.sex or "",
            r.facility.name if r.facility else "",
            r.case_manager or "",
            r.next_appointment.strftime("%Y-%m-%d") if r.next_appointment else "",
            r.vl_due_date.strftime("%Y-%m-%d") if r.vl_due_date else "",
            r.vl_status_display,
            "Yes" if r.vl_eligible else "No",
            suppression,
            "Yes" if r.eac else "No",
            "Eligible" if r.ahd else "Not Eligible",
            r.tpt_status or "",
            r.get_tb_status_display() if hasattr(r, "get_tb_status_display") else "",
            r.tracking_date_1.strftime("%Y-%m-%d") if r.tracking_date_1 else "",
            r.tracking_date_2.strftime("%Y-%m-%d") if r.tracking_date_2 else "",
            r.tracking_date_3.strftime("%Y-%m-%d") if r.tracking_date_3 else "",
            r.tracked_by or "",
            r.missed_reason or "",
            discontinued,
            r.discontinued_reason or "",
            r.discontinued_date.strftime("%Y-%m-%d") if r.discontinued_date else "",
            r.returned_date.strftime("%Y-%m-%d") if r.returned_date else "",
            r.remark or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="Expected_Refills_{today}.xlsx"'

    wb.save(response)
    return response
# ================= REFILL CREATE =================



@login_required
def refill_create_or_update(request, pk=None):

    today = timezone.now().date()
    refill = get_object_or_404(Refill, pk=pk) if pk else Refill()

    form = RefillForm(request.POST or None, instance=refill)

    if request.method == "POST":
        if form.is_valid():
            refill = form.save(commit=False)

            if refill.last_pickup_date and refill.months_of_refill_days:
                days = int(float(refill.months_of_refill_days) * 30)
                refill.next_appointment = refill.last_pickup_date + timedelta(days=days)
            else:
                refill.next_appointment = None

            refill.missed_appointment = (
                refill.next_appointment < today if refill.next_appointment else False
            )

            if refill.tpt_start_date:
                refill.tpt_expected_completion = refill.tpt_start_date + timedelta(days=180)

            refill.save()

            messages.success(request, "Refill record saved successfully.")
            return redirect("track_refills")

    eligible, status = attach_vl_status(refill)

    return render(request, "refill_form.html", {
        "form": form,
        "today": today,
        "vl_eligible": eligible,
        "vl_status": status,
    })


# ==========================================================
# ADD / UPDATE USING UNIQUE ID
# ==========================================================
@login_required
def refill_add_or_update(request, unique_id=None):

    refill = (
        Refill.objects.filter(unique_id=unique_id)
        .order_by("-last_pickup_date")
        .first()
        if unique_id else None
    )

    today = timezone.now().date()

    form = RefillForm(request.POST or None, instance=refill)

    latest_tb = getattr(refill, "latest_tb", None)

    if request.method == "POST":
        if form.is_valid():
            refill_instance = form.save(commit=False)

            if refill_instance.last_pickup_date and refill_instance.months_of_refill_days:
                days = int(float(refill_instance.months_of_refill_days) * 30)
                refill_instance.next_appointment = refill_instance.last_pickup_date + timedelta(days=days)

            refill_instance.missed_appointment = (
                refill_instance.next_appointment < today if refill_instance.next_appointment else False
            )

            if refill_instance.tpt_start_date:
                refill_instance.tpt_expected_completion = refill_instance.tpt_start_date + timedelta(days=180)

            refill_instance.save()
            messages.success(request, "Refill record saved successfully.")
            return redirect("track_refills")

    eligible, status = attach_vl_status(refill)

    return render(request, "refill_form.html", {
        "form": form,
        "today": today,
        "latest_tb": latest_tb,
        "vl_eligible": eligible,
        "vl_status": status,
    })

# ==========================================================
# CREATE REFILL
# ==========================================================


@login_required
def refill_create(request, unique_id=None):

    refill = Refill.objects.filter(unique_id=unique_id).first() if unique_id else None
    today = timezone.now().date()

    form = RefillForm(request.POST or None, instance=refill)

    if request.method == "POST":
        if form.is_valid():
            refill_obj = form.save(commit=False)

            if refill_obj.last_pickup_date and refill_obj.months_of_refill_days:
                days = int(float(refill_obj.months_of_refill_days) * 30)
                refill_obj.next_appointment = refill_obj.last_pickup_date + timedelta(days=days)

            refill_obj.missed_appointment = (
                refill_obj.next_appointment < today if refill_obj.next_appointment else False
            )

            if refill_obj.tpt_start_date:
                refill_obj.tpt_expected_completion = refill_obj.tpt_start_date + timedelta(days=180)

            refill_obj.save()

            messages.success(request, "Refill record saved successfully.")
            return redirect("track_refills")

    eligible, status = attach_vl_status(refill)

    return render(request, "refill_form.html", {
        "form": form,
        "today": today,
        "vl_eligible": eligible,
        "vl_status": status,
    })

# ==========================================================
# UPDATE REFILL
# ==========================================================
@login_required
def refill_update(request, pk):

    refill = get_object_or_404(Refill, pk=pk)
    today = timezone.now().date()

    form = RefillForm(request.POST or None, instance=refill)

    if request.method == "POST":
        if form.is_valid():
            refill = form.save(commit=False)

            if refill.last_pickup_date and refill.months_of_refill_days:
                days = int(float(refill.months_of_refill_days) * 30)
                refill.next_appointment = refill.last_pickup_date + timedelta(days=days)

            refill.missed_appointment = (
                refill.next_appointment < today if refill.next_appointment else False
            )

            if refill.tpt_start_date:
                refill.tpt_expected_completion = refill.tpt_start_date + timedelta(days=180)

            refill.save()

            messages.success(request, "Refill updated successfully.")
            return redirect("track_refills")

    eligible, status = attach_vl_status(refill)

    return render(request, "refill_form.html", {
        "form": form,
        "today": today,
        "vl_eligible": eligible,
        "vl_status": status,
    })
    
    
    
    
    
    
@login_required
def track_refills(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    def parse_date(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date()
        except:
            return None

    start_date = parse_date(start_date)
    end_date = parse_date(end_date)

    # ================= BASE QUERY =================
    user = request.user

    if user.is_superuser:
        # SUPERUSER: sees all data + all facilities for filtering
        refills = Refill.objects.select_related("facility").all()
        facilities = Facility.objects.all().order_by("name")
    else:
        facility = get_user_facility(user)
        if not facility:
            return redirect("login")

        # NORMAL USER: restricted data
        refills = Refill.objects.select_related("facility").filter(facility=facility)
        facilities = Facility.objects.filter(id=facility.id)

    # ================= FILTER: FACILITY (SUPERUSER ONLY) =================
    if facility_id and user.is_superuser:
        refills = refills.filter(facility_id=facility_id)

    # ================= CASE MANAGERS =================
    case_managers_qs = refills.exclude(
        case_manager__isnull=True
    ).exclude(
        case_manager__exact=""
    ).values_list("case_manager", flat=True).distinct()

    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm and cm.strip()})

    # ================= DATE RANGE =================
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    if start_date and end_date:
        refills = refills.filter(last_pickup_date__range=[start_date, end_date])
    elif start_date:
        refills = refills.filter(last_pickup_date__gte=start_date)
    elif end_date:
        refills = refills.filter(last_pickup_date__lte=end_date)
    else:
        refills = refills.filter(last_pickup_date__range=[month_start, month_end])

    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    # ================= ENRICH =================
    for r in refills:
        r.days_missed_display = r.days_missed or 0
        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status

        if not r.vl_eligible:
            if r.patient_discontinued == "Y":
                r.vl_status_display = "Not Eligible (Discontinued)"
            elif r.days_missed_display >= 28:
                r.vl_status_display = "Not Eligible (IIT)"

    # ================= PAGINATION =================
    paginator = Paginator(refills, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "track_refills.html", {
        "facilities": facilities,
        "case_managers": case_managers,
        "selected_facility": facility_id,
        "selected_case_manager": selected_case_manager,
        "today": today,
        "page_obj": page_obj,
    })
    
    
    
# ================= EXPORT =================
@login_required
def export_track_refills_view(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    def parse_date(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date()
        except:
            return None

    start_date = parse_date(start_date)
    end_date = parse_date(end_date)

    user = request.user

    if user.is_superuser:
        refills = Refill.objects.select_related("facility").all()
    else:
        facility = get_user_facility(user)
        if not facility:
            return redirect("login")

        refills = Refill.objects.select_related("facility").filter(facility=facility)

    # ================= FILTER LOGIC =================
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    if start_date and end_date:
        refills = refills.filter(last_pickup_date__range=[start_date, end_date])
    elif start_date:
        refills = refills.filter(last_pickup_date__gte=start_date)
    elif end_date:
        refills = refills.filter(last_pickup_date__lte=end_date)
    else:
        refills = refills.filter(last_pickup_date__range=[month_start, month_end])

    if facility_id and user.is_superuser:
        refills = refills.filter(facility_id=facility_id)

    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    # ================= EXCEL =================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Track Refills Data"

    headers = [
        "Unique ID", "Facility", "Last Pickup Date", "Refill Days",
        "Sex", "Current Regimen", "Case Manager", "Next Appointment",
        "Days Missed", "VL Status",
        "Tracking Date 1", "Tracking Date 2", "Tracking Date 3",
        "Tracked By", "Missed Reason",
        "Patient Discontinued", "Discontinued Reason",
        "Discontinued Date", "Returned Date",
    ]

    ws.append(headers)

    for r in refills:

        r.calculate_dates()

        next_appointment = r.next_appointment
        last_pickup = r.last_pickup_date

        days_missed = (
            (today - next_appointment).days
            if next_appointment and next_appointment < today
            else 0
        )

        vl_status = r.vl_status

        if not r.is_vl_eligible_program:
            if r.patient_discontinued == "Y":
                vl_status = "Not Eligible (Discontinued)"
            elif days_missed >= 28:
                vl_status = "Not Eligible (IIT)"

        ws.append([
            r.unique_id,
            r.facility.name if r.facility else "",
            last_pickup.strftime("%Y-%m-%d") if last_pickup else "",
            r.months_of_refill_days,
            r.sex,
            r.current_regimen,
            r.case_manager or "",
            next_appointment.strftime("%Y-%m-%d") if next_appointment else "",
            days_missed,
            vl_status,
            r.tracking_date_1.strftime("%Y-%m-%d") if r.tracking_date_1 else "",
            r.tracking_date_2.strftime("%Y-%m-%d") if r.tracking_date_2 else "",
            r.tracking_date_3.strftime("%Y-%m-%d") if r.tracking_date_3 else "",
            r.tracked_by or "",
            r.missed_reason or "",
            "Yes" if r.patient_discontinued == "Y" else "No",
            r.discontinued_reason or "",
            r.discontinued_date.strftime("%Y-%m-%d") if r.discontinued_date else "",
            r.returned_date.strftime("%Y-%m-%d") if r.returned_date else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="Track_Refills_{today}.xlsx"'

    wb.save(response)
    return response

@login_required
def daily_refill_list(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")

    user = request.user

    # ================= BASE QUERY =================
    if user.is_superuser:

        refills = Refill.objects.filter(
            next_appointment=today
        ).select_related("facility")

        facilities = Facility.objects.all().order_by("name")

        # optional filter for superuser
        if facility_id:
            refills = refills.filter(facility_id=facility_id)

    else:
        facility = get_user_facility(user)
        if not facility:
            return redirect("login")

        refills = Refill.objects.filter(
            next_appointment=today,
            facility=facility
        ).select_related("facility")

        facilities = [facility]

    # ================= CASE MANAGER FILTER =================
    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    # ================= CASE MANAGERS =================
    case_managers_qs = refills.exclude(
        case_manager__isnull=True
    ).exclude(
        case_manager__exact=""
    ).values_list("case_manager", flat=True).distinct()

    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm})

    # ================= PROCESS =================
    for r in refills:

        r.age_display = r.age or "Unknown"
        r.tb_status_display = r.tb_status or "Not Screened"
        r.days_missed_display = r.days_missed or 0

        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status
        r.vl_due_date_display = r.vl_due_date

        r.eac_status_display = r.eac_status
        r.ahd_display = "Eligible" if r.ahd else "Not Eligible"

    return render(request, "daily_refill_list.html", {
        "facilities": facilities,
        "selected_facility": facility_id,
        "case_managers": case_managers,
        "selected_case_manager": selected_case_manager,
        "today": today,
        "refills": refills,
    })
    
    
    

# ================= EXPORT MISSED REFILLS =================
@login_required
def export_missed_refills_view(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    case_manager = request.GET.get("case_manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    search_unique_id = request.GET.get("search_unique_id")

    user = request.user

    # ================= BASE QUERY =================
    if user.is_superuser:
        refills = Refill.objects.select_related("facility").all()
    else:
        facility = get_user_facility(user)
        if not facility:
            return redirect("login")

        refills = Refill.objects.select_related("facility").filter(facility=facility)

    # ================= FILTERS =================
    if facility_id and user.is_superuser:
        refills = refills.filter(facility_id=facility_id)

    if case_manager:
        refills = refills.filter(case_manager__iexact=case_manager.strip())

    if search_unique_id:
        refills = refills.filter(unique_id__icontains=search_unique_id.strip())

    if start_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__gte=start_date)
        except:
            pass

    if end_date:
        try:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__lte=end_date)
        except:
            pass

    # ================= MISSED ONLY =================
    missed_list = refills.filter(
        next_appointment__lt=today,
        next_appointment__isnull=False
    ).order_by("next_appointment")

    # ================= EXPORT =================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missed Refills Data"

    headers = [
        'Unique ID',
        'Facility',
        'Sex',
        'Current Regimen',
        'Case Manager',
        'Last Pickup Date',
        'Next Appointment',
        'Days Missed',
        'IIT Status',
        'VL Status',
        'Tracking Date 1',
        'Tracking Date 2',
        'Tracking Date 3',
        'Tracked By',
        'Patient Discontinued',
        'Discontinued Reason',
        'Discontinued Date',
        'Returned Date',
    ]

    ws.append(headers)

    for r in missed_list:

        days_missed = r.days_missed or 0
        iit_status = r.iit_status

        vl_eligible = r.is_vl_eligible_program
        is_suppressed = r.is_suppressed

        if not vl_eligible:
            vl_status = "Not Eligible"

        else:
            if not r.vl_sample_collection_date:
                vl_status = "Eligible (No VL Yet)"
            elif r.vl_result is None:
                vl_status = "VL Pending"
            elif is_suppressed:
                vl_status = "Suppressed (<1000)"
            else:
                vl_status = "Not Suppressed (≥1000)"

        ws.append([
            r.unique_id,
            r.facility.name if r.facility else "",
            r.sex,
            r.current_regimen,
            r.case_manager or "",
            r.last_pickup_date.strftime("%Y-%m-%d") if r.last_pickup_date else "",
            r.next_appointment.strftime("%Y-%m-%d") if r.next_appointment else "",
            days_missed,
            iit_status,
            vl_status,

            r.tracking_date_1.strftime("%Y-%m-%d") if r.tracking_date_1 else "",
            r.tracking_date_2.strftime("%Y-%m-%d") if r.tracking_date_2 else "",
            r.tracking_date_3.strftime("%Y-%m-%d") if r.tracking_date_3 else "",
            r.tracked_by or "",

            dict(r._meta.get_field('patient_discontinued').choices).get(
                r.patient_discontinued, ""
            ) if r.patient_discontinued else "",

            r.discontinued_reason or "",
            r.discontinued_date.strftime("%Y-%m-%d") if r.discontinued_date else "",
            r.returned_date.strftime("%Y-%m-%d") if r.returned_date else "",
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="Missed_Refills_{today}.xlsx"'

    wb.save(response)
    return response

@login_required
def missed_refills(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    case_manager = request.GET.get("case_manager")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    search_unique_id = request.GET.get("search_unique_id")

    # ================= BASE QUERYSET =================
    if request.user.is_superuser:
        refills = Refill.objects.all().select_related("facility")
        facilities = Facility.objects.all().order_by("name")

        # optional filter (superuser only)
        if facility_id:
            refills = refills.filter(facility_id=facility_id)

    else:
        refills = Refill.objects.filter(
            facility__memberships__user=request.user
        ).select_related("facility")

        facility = get_user_facility(request.user)
        facilities = [facility] if facility else []

    # ================= FILTERS =================
    if case_manager:
        refills = refills.filter(case_manager__iexact=case_manager.strip())

    if search_unique_id:
        refills = refills.filter(unique_id__icontains=search_unique_id.strip())

    if start_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__gte=start_date)
        except:
            pass

    if end_date:
        try:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            refills = refills.filter(next_appointment__lte=end_date)
        except:
            pass

    # ================= MISSED ONLY =================
    missed_list = refills.filter(
        next_appointment__lt=today,
        next_appointment__isnull=False
    ).order_by("next_appointment")

    # ================= ENRICH =================
    for r in missed_list:

        r.days_missed_display = r.days_missed or 0
        r.iit_status_display = r.iit_status

        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status
        r.vl_due_date_display = r.vl_due_date

        if not r.vl_sample_collection_date:
            r.vl_result_display = "No VL Done"

        elif r.vl_result is None:
            r.vl_result_display = "Pending"

        elif r.is_suppressed:
            r.vl_result_display = "Suppressed (<1000)"

        else:
            r.vl_result_display = "Not Suppressed (≥1000)"

        r.tracking_1 = r.tracking_date_1
        r.tracking_2 = r.tracking_date_2
        r.tracking_3 = r.tracking_date_3
        r.tracked_by_display = r.tracked_by

        r.discontinued_display = dict(
            r._meta.get_field('patient_discontinued').choices
        ).get(r.patient_discontinued, "") if r.patient_discontinued else ""

    # ================= PAGINATION =================
    paginator = Paginator(missed_list, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ================= CASE MANAGERS =================
    case_managers_qs = refills.exclude(
        case_manager__isnull=True
    ).exclude(
        case_manager__exact=""
    ).values_list("case_manager", flat=True).distinct()

    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm})

    query_params = request.GET.copy()
    query_params.pop("page", None)

    return render(request, "missed_refills.html", {
        "page_obj": page_obj,
        "today": today,
        "total_missed": missed_list.count(),
        "facilities": facilities,
        "case_managers": case_managers,
        "selected_facility": facility_id,
        "selected_case_manager": case_manager,
        "selected_start_date": start_date,
        "selected_end_date": end_date,
        "search_unique_id": search_unique_id,
        "query_params": query_params.urlencode(),
    })
    
    
@login_required
def track_vl(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")

    user = request.user

    # ================= BASE QUERY =================
    if user.is_superuser:

        refills = Refill.objects.select_related("facility").all()
        facilities = Facility.objects.all().order_by("name")

        # optional filter for superuser
        if facility_id:
            refills = refills.filter(facility_id=facility_id)

    else:
        refills = Refill.objects.filter(
            facility__memberships__user=user
        ).select_related("facility")

        facility = get_user_facility(user)
        facilities = [facility] if facility else []

    # ================= CASE MANAGERS =================
    case_managers_qs = refills.exclude(
        case_manager__isnull=True
    ).exclude(
        case_manager__exact=""
    ).values_list("case_manager", flat=True).distinct()

    case_managers = sorted({cm.strip() for cm in case_managers_qs if cm and cm.strip()})

    # ================= FILTER: CASE MANAGER =================
    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    # ================= PROCESS =================
    processed_refills = []

    for r in refills:

        # ================= VL ENGINE =================
        r.vl_eligible = r.is_vl_eligible_program
        r.vl_status_display = r.vl_status
        r.vl_due_date_display = r.vl_due_date
        r.days_missed_display = r.days_missed

        # ================= SUPPRESSION =================
        suppressed = r.is_suppressed

        if suppressed is None:
            r.suppression_display = "No Result"
        elif suppressed:
            r.suppression_display = "Suppressed (<1000)"
        else:
            r.suppression_display = "Not Suppressed (≥1000)"

        # ================= OVERRIDE LOGIC =================
        if not r.vl_eligible:
            if r.patient_discontinued == "Y":
                r.vl_status_display = "Not Eligible (Discontinued)"
            elif (r.days_missed or 0) >= 28:
                r.vl_status_display = "Not Eligible (IIT)"

        processed_refills.append(r)

    # ================= SORT =================
    processed_refills.sort(
        key=lambda x: (
            not x.vl_eligible,
            x.vl_due_date or today
        )
    )

    # ================= PAGINATION =================
    paginator = Paginator(processed_refills, 10)
    page_number = request.GET.get("page")
    vl_refills = paginator.get_page(page_number)

    return render(request, "track_vl.html", {
        "facilities": facilities,
        "selected_facility": facility_id,
        "selected_case_manager": selected_case_manager,
        "case_managers": case_managers,
        "vl_refills": vl_refills,
        "today": today,
    })
    
    
    
    
    
@login_required
def export_vl_view(request):

    today = timezone.now().date()

    facility_id = request.GET.get("facility")
    selected_case_manager = request.GET.get("case_manager")

    # ================= BASE QUERY =================
    if request.user.is_superuser:
        refills = Refill.objects.select_related("facility")
    else:
        refills = Refill.objects.filter(
            facility__memberships__user=request.user
        ).select_related("facility")

    # ================= FILTERS =================
    if facility_id:
        refills = refills.filter(facility_id=facility_id)

    if selected_case_manager:
        refills = refills.filter(case_manager__iexact=selected_case_manager.strip())

    refills = refills.order_by("-vl_sample_collection_date")

    # ================= EXCEL =================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Track VL"

    headers = [
        "Unique ID",
        "Facility",
        "VL Sample Collection Date",
        "Last Pickup Date",
        "Case Manager",
        "Days Missed",
        "VL Status",
        "VL Due Date",

        "Tracking Date 1",
        "Tracking Date 2",
        "Tracking Date 3",
        "Tracked By",

        "Patient Discontinued",
        "Discontinued Reason",
        "Discontinued Date",
        "Returned Date",
    ]

    ws.append(headers)

    # ================= ROWS =================
    for r in refills:

        days_missed = r.days_missed or 0

        vl_status = r.vl_status
        vl_due_date = r.vl_due_date
        vl_eligible = r.is_vl_eligible_program

        if not vl_eligible:
            if r.patient_discontinued == "Y":
                vl_status = "Not Eligible (Discontinued)"
            elif days_missed >= 28:
                vl_status = "Not Eligible (IIT)"

        ws.append([
            r.unique_id,
            r.facility.name if r.facility else "",
            r.vl_sample_collection_date.strftime("%Y-%m-%d") if r.vl_sample_collection_date else "",
            r.last_pickup_date.strftime("%Y-%m-%d") if r.last_pickup_date else "",
            r.case_manager or "",
            days_missed,
            vl_status,
            vl_due_date.strftime("%Y-%m-%d") if vl_due_date else "",

            r.tracking_date_1.strftime("%Y-%m-%d") if r.tracking_date_1 else "",
            r.tracking_date_2.strftime("%Y-%m-%d") if r.tracking_date_2 else "",
            r.tracking_date_3.strftime("%Y-%m-%d") if r.tracking_date_3 else "",
            r.tracked_by or "",

            r.patient_discontinued or "",
            r.discontinued_reason or "",
            r.discontinued_date.strftime("%Y-%m-%d") if r.discontinued_date else "",
            r.returned_date.strftime("%Y-%m-%d") if r.returned_date else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="Track_VL_{today}.xlsx"'

    wb.save(response)
    return response